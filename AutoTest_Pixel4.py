import pyautogui
import pyscreenshot as ImageGrab
import time
from PIL import ImageGrab
from functools import partial

import openpyxl
from openpyxl.styles import Font, PatternFill       # 載入 Font 和 PatternFill 模組
import datetime

import cv2
import numpy as np
from matplotlib import pyplot as plt
import pytesseract

def checkExits(PicName, Times = 30, findGameIcon = False):
    nCheckTimes = 0
    iconAppCheck = False
    PicNamePath = 'D:/Selenium/auto_test/Pixel4_Pic/' + PicName + '.png'
    imgIcon = cv2.imread(PicNamePath)
    print(PicNamePath)
    iconW = imgIcon.shape[1]
    iconH = imgIcon.shape[0]

    while iconAppCheck == False:
        screen = pyautogui.screenshot()
        screen.save("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        screen = cv2.imread("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        res = cv2.matchTemplate(screen,imgIcon,1)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        #print(min_val, max_val, min_loc, max_loc)
        top_left = min_loc
        #print(top_left)

        bottom_right = (top_left[0] + iconW, top_left[1] + iconH)
        imgscreenIcon = ImageGrab.grab(bbox=(top_left[0], top_left[1], bottom_right[0], bottom_right[1]))
        #存檔確認比對圖案是否正確,此問題待解，希望可以不要再存檔重新讀取
        PicIconTempName = "D:/Selenium/auto_test/PicTemp/iconTemp.png"
        imgscreenIcon.save(PicIconTempName)
        imgscreenIcon2 = cv2.imread(PicIconTempName)
        hsv_base = cv2.cvtColor(imgIcon, cv2.COLOR_BGR2HSV)
        hsv_test = cv2.cvtColor(imgscreenIcon2, cv2.COLOR_BGR2HSV)
        h_bins = 50
        s_bins = 60
        histSize = [h_bins, s_bins]
        h_ranges = [0, 180]
        s_ranges = [0, 256]
        ranges = h_ranges + s_ranges
        
        channels = [0, 1]
        hist_base = cv2.calcHist([hsv_base], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_base, hist_base, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        hist_test = cv2.calcHist([hsv_test], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_test, hist_test, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        
        compare_method = cv2.HISTCMP_CORREL
        base_test = cv2.compareHist(hist_base, hist_test, compare_method)

        print(PicName + ' Similarity = ', base_test)
        if base_test > 0.9:
            print('Found ' + PicName)
            pyautogui.moveTo(top_left[0] + iconW/2,top_left[1] + iconH/2,0.5)
            pyautogui.click()
            iconAppCheck = True
        else:
            if findGameIcon == True:
                pyautogui.moveTo(3857,1241,0.5)
                pyautogui.mouseDown(3857,1241)
                pyautogui.moveTo(3857,748,2)
                time.sleep(0.3)
                pyautogui.mouseUp(3857,748)
            print(nCheckTimes)
            nCheckTimes = nCheckTimes + 1
            time.sleep(1)
            if nCheckTimes >= Times:
                print('compare over times ' + PicName)
                return False
    return True

def checkTableExits(tableNum,clickX,clickY):
    time.sleep(2)
    PicNamePath = 'D:/Selenium/auto_test/Pixel4_Pic/TableClickCompare.png'
    imgIcon = cv2.imread(PicNamePath)

    iconW = imgIcon.shape[1]
    iconH = imgIcon.shape[0]
    screen = pyautogui.screenshot()
    screen.save("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
    screen = cv2.imread("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
    res = cv2.matchTemplate(screen,imgIcon,1)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
    #print(min_val, max_val, min_loc, max_loc)
    top_left = min_loc
    #print(top_left)

    bottom_right = (top_left[0] + iconW, top_left[1] + iconH)
    pyautogui.moveTo(clickX,clickY,0.2)
    pyautogui.click()
    time.sleep(0.1)
    imgscreenIcon = ImageGrab.grab(bbox=(top_left[0], top_left[1], bottom_right[0], bottom_right[1]))
    #存檔確認比對圖案是否正確,此問題待解，希望可以不要再存檔重新讀取
    PicIconTempName = "D:/Selenium/auto_test/PicTemp/iconTemp.png"
    imgscreenIcon.save(PicIconTempName)
    imgscreenIcon2 = cv2.imread(PicIconTempName)
    hsv_base = cv2.cvtColor(imgIcon, cv2.COLOR_BGR2HSV)
    hsv_test = cv2.cvtColor(imgscreenIcon2, cv2.COLOR_BGR2HSV)
    h_bins = 50
    s_bins = 60
    histSize = [h_bins, s_bins]
    h_ranges = [0, 180]
    s_ranges = [0, 256]
    ranges = h_ranges + s_ranges
        
    channels = [0, 1]
    hist_base = cv2.calcHist([hsv_base], channels, None, histSize, ranges, accumulate=False)
    cv2.normalize(hist_base, hist_base, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
    hist_test = cv2.calcHist([hsv_test], channels, None, histSize, ranges, accumulate=False)
    cv2.normalize(hist_test, hist_test, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        
    compare_method = cv2.HISTCMP_CORREL
    base_test = cv2.compareHist(hist_base, hist_test, compare_method)

    print('Table '+ str(tableNum) + ' Similarity = ', base_test)
    if base_test < 0.9:
        sheet.cell(nGameCnt + 2,tableNum + 2).value = "OK"
        print("table " + str(tableNum) + " actived!!!!")
    else:
        sheet.cell(nGameCnt + 2,tableNum + 2).value = "X"
        sheet.cell(nGameCnt + 2,tableNum + 2).fill = PatternFill(fill_type="solid", fgColor="FF0000")
        print("table " + str(tableNum) + " untouchedble~~~")
        return False
    return True

def checkDeskExits():
    #第 1 排
    tableCnt = 1
    checkTableExits(tableCnt,3629,426)
    
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3745,426)
    
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3857,426)
    
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3973,426)
    
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,4086,426)

    #第 2 排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3629,570)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3745,570)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3857,570)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3973,570)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,4086,570)

    #第 3 排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3629,712)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3745,712)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3857,712)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3973,712)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,4086,712)


    #第 4 排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3629,860)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3745,860)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3857,860)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3973,860)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,4086,860)

    #第 5 排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3629,1001)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3745,1001)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3857,1001)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3973,1001)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,4086,1001)

    #第 6 排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3629,1147)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3745,1147)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3857,1147)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3973,1147)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,4086,1147)
    return

def checkGameExits(GameName):
    #檢查是否到遊戲館大廳
    if checkExits('Trial_MenuMore') == False:
        print('Trial_MenuMore Fail')
    #關閉選單
    pyautogui.click()
    
    #檢查 game Icon是否存在
    if checkExits(GameName,20,True) == False:
        print(GameName + ' check Fail')

    #檢查進入遊戲按鈕是否存在
    if checkExits('IntoDesk') == False:
        print('IntoDesk Fail')

    #檢查 遊戲內橫幅圖是否存在，用以確認真的進到遊戲選桌畫面中
    #gameBaner = GameName + '_Banner'
    #if checkExits(gameBaner) == False:
    #    print(gameBaner +' check Fail')

    #檢查桌詳細資訊按鈕是否存在，順便確認是否真的進到遊戲選桌畫面中
    if checkExits('TableDetail') == False:
        print('TableDetail check Fail')

    #檢查桌詳細資訊是否可以打開，順便確認是否真的進到遊戲選桌畫面中
    if checkExits('TableDetailInfo') == True:
        sheet.cell(nGameCnt + 2,2).value = 'OK'
    else:
        sheet.cell(nGameCnt + 2,2).value = 'Fail'
        print('TableDetailInfo check Fail')
    
    #關閉桌詳細資訊
    if checkExits('TableDetail') == False:
        print('TableDetail check Fail')


    sheet.cell(nGameCnt + 2,1).value = GameName
    #檢查 遊戲內 各桌是否可以按
    checkDeskExits()
    
    wb.save("D:/GoogleDrive/RegressionTest/RegressionTest.xlsx")


    #檢查離開遊戲按鈕是否存在
    if checkExits('BacktoGameLobby') == False:
        print('BacktoGameLobby Fail')

    return

wb = openpyxl.load_workbook('D:/GoogleDrive/RegressionTest/RT_DeskCheck.xlsx')
localtime = time.localtime()
sheet = wb.create_sheet(str(time.strftime("Pixel4_%y%m%d-%H-%M", localtime)),0)
ImageGrab.grab = partial(ImageGrab.grab, all_screens=True)

#夜神 ForAutoTest
#pyautogui.moveTo(4752,56,0.5)
#pyautogui.doubleClick()

#檢查模擬器桌面icon是否存在，若是則點擊進去，若否，則等待，預設等待超過10次則代表失敗。
#if checkExits('NOX_HomeWudiAppIcon') == False:
#    print('NOX_HomeWudiAppIcon Fail')

#檢查會員登入按鈕是否存在
if checkExits('memberLoginIcon', 30) == False:
    print('memberLoginIcon Fail')

#檢查會員帳號輸入空間是否存在
if checkExits('MemberAccount', 30) == False:
    print('MemberAccount Fail')
time.sleep(1)
pyautogui.typewrite(['Q','C','T','E','S','T','0','1'],'0.3')

#檢查會員帳號輸入空間是否存在
if checkExits('MemberPassword', 30) == False:
    print('MemberPassword Fail')
time.sleep(1)
pyautogui.typewrite(['Q','q','1','1','1','1','1','1'],'0.3')

#檢查公告關閉按鈕是否存在
if checkExits('AccPWConfirm') == False:
    print('AccPWConfirm Fail')

#檢查廣告關閉按鈕是否存在
if checkExits('AdClose') == False:
    print('AdClose Fail')

#檢查廣告關閉按鈕是否存在
if checkExits('Announce') == False:
    print('Announce Fail')

#檢查電子館大廳按鈕是否存在
if checkExits('GoInToGameLobby') == False:
    print('GoInToGameLobby Fail')

#檢查進入遊戲按鈕是否存在
if checkExits('TransferMoney') == False:
    print('TransferMoney Fail')

nGameCnt = 0
checkGameExits('6001')

nGameCnt = nGameCnt + 1
checkGameExits('6002')

nGameCnt = nGameCnt + 1
checkGameExits('6003')

nGameCnt = nGameCnt + 1
checkGameExits('6004')

nGameCnt = nGameCnt + 1
checkGameExits('6005')

nGameCnt = nGameCnt + 1
checkGameExits('6006')

nGameCnt = nGameCnt + 1
checkGameExits('6008')

nGameCnt = nGameCnt + 1
checkGameExits('6009')

nGameCnt = nGameCnt + 1
checkGameExits('6010')

nGameCnt = nGameCnt + 1
checkGameExits('6012')

nGameCnt = nGameCnt + 1
checkGameExits('6013')

nGameCnt = nGameCnt + 1
checkGameExits('6014')

nGameCnt = nGameCnt + 1
checkGameExits('6015')

nGameCnt = nGameCnt + 1
checkGameExits('6016')

nGameCnt = nGameCnt + 1
checkGameExits('6017')