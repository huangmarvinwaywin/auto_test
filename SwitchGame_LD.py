import pyautogui
import pyscreenshot as ImageGrab
import time
from PIL import ImageGrab
from functools import partial
from PIL import Image

import openpyxl
from openpyxl.styles import Font, PatternFill       # 載入 Font 和 PatternFill 模組
import datetime

import cv2
import numpy as np
from matplotlib import pyplot as plt
import imagehash
nGameCnt = 0


def checkExits(PicName,bClickIconFound = True, Times = 15, findGameIcon = False):
    nCheckTimes = 0
    iconAppCheck = False
    PicNamePath = 'D:/Selenium/auto_test/SwitchGame/' + PicName + '.png'
    imgIcon = cv2.imread(PicNamePath)
    print(PicNamePath)
    iconW = imgIcon.shape[1]
    iconH = imgIcon.shape[0]

    while iconAppCheck == False:
        screen = pyautogui.screenshot()
        screen.save("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        screen = cv2.imread("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        res = cv2.matchTemplate(screen,imgIcon,1)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        #print(min_val, max_val, min_loc, max_loc)
        top_left = min_loc
        #print(top_left)

        bottom_right = (top_left[0] + iconW, top_left[1] + iconH)
        imgscreenIcon = ImageGrab.grab(bbox=(top_left[0], top_left[1], bottom_right[0], bottom_right[1]))
        #存檔確認比對圖案是否正確,此問題待解，希望可以不要再存檔重新讀取
        PicIconTempName = "D:/Selenium/auto_test/PicTemp/iconTemp.png"
        imgscreenIcon.save(PicIconTempName)
        imgscreenIcon2 = cv2.imread(PicIconTempName)
        hsv_base = cv2.cvtColor(imgIcon, cv2.COLOR_BGR2HSV)
        hsv_test = cv2.cvtColor(imgscreenIcon2, cv2.COLOR_BGR2HSV)
        h_bins = 50
        s_bins = 60
        histSize = [h_bins, s_bins]
        h_ranges = [0, 180]
        s_ranges = [0, 256]
        ranges = h_ranges + s_ranges
        
        channels = [0, 1]
        hist_base = cv2.calcHist([hsv_base], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_base, hist_base, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        hist_test = cv2.calcHist([hsv_test], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_test, hist_test, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        
        compare_method = cv2.HISTCMP_CORREL
        base_test = cv2.compareHist(hist_base, hist_test, compare_method)

        print(PicName + ' Similarity = ', base_test)
        if base_test > 0.9:
            print('Found ' + PicName)
            if bClickIconFound == True:
                pyautogui.moveTo(top_left[0] + iconW/2,top_left[1] + iconH/2,0.5)
                pyautogui.click()
            iconAppCheck = True
        else:
            if findGameIcon == True:
                pyautogui.moveTo(2850,1100,0.5)
                pyautogui.mouseDown(2850,1100)
                pyautogui.moveTo(2850,650,2)
                time.sleep(0.3)
                pyautogui.mouseUp(2850,650)
            print(nCheckTimes)
            nCheckTimes = nCheckTimes + 1
            time.sleep(1)
            if nCheckTimes >= Times:
                print('compare over times ' + PicName)
                return False
    return True

def SwitchToGame(GameName,WritePos,EnterMode,SwitchMode):
    
    #EnterMode 0->ToDesk;1->ToGame
    #檢查是否到遊戲館大廳
    if checkExits('Trial_MenuMore',False) == False:
        print('Trial_MenuMore Fail')
    
    time.sleep(1)


    #檢查 game Icon是否存在
    if checkExits(GameName,True,20,True) == False:
        print(GameName + ' check Fail')

    EnterGame = 'D:/Selenium/auto_test/SwitchGame/' + GameName + '_Game.png'
    if EnterMode == 0:
        #進入選桌
        if checkExits('IntoDesk') == False:
            print('IntoDesk Fail')
    else:
        if checkExits(EnterGame) == False:
            print('EnterGame Fail')
    
    if checkExits('SwitchTo') == False:
        print('SwitchTo Fail')
    #Switch mode == 0, 老虎機->老虎機    
    #Switch mode == 1, 老虎機->特色
    #Switch mode == 2, 特色->特色
    #Switch mode == 3, 特色->老虎機
    if SwitchMode == 0:
        if checkExits('SP_Game') == False:
            print('SP_Game Fail')
        time.sleep(0.2)
        if checkExits('Slot_Game') == False:
            print('Slot_Game Fail')
    elif SwitchMode == 1:
        if checkExits('SP_Game') == False:
            print('Slot_Game Fail')
    elif SwitchMode == 2:
        if checkExits('Slot_Game') == False:
            print('Slot_Game Fail')
        time.sleep(0.2)
        if checkExits('SP_Game') == False:
            print('SP_Game Fail')
    else :
        if checkExits('Slot_Game') == False:
            print('Slot_Game Fail')


 
    #sheet.cell(WritePos + 2,1).value = GameName
    #wb.save("D:/GoogleDrive/RegressionTest/RT_SwitchGame.xlsx")


    #檢查是否到遊戲館大廳
    #if checkExits('CheckInDesk',False) == False:
    #    print('CheckInDesk Fail')


    return True

   
def main():

    sheet.cell(1,1).value = "Game"
    sheet.cell(1,2).value = "Desk to Desk"
    sheet.cell(1,3).value = "Desk to Game"
    sheet.cell(1,4).value = "Game to Desk"
    sheet.cell(1,5).value = "Game to Game"

    wb.save("D:/GoogleDrive/RegressionTest/RT_SwitchGame.xlsx")

    #雷電 L_ForAutoTest
    pyautogui.moveTo(4191,46,0.5)
    pyautogui.doubleClick()
    
    #檢查模擬器桌面icon是否存在，若是則點擊進去，若否，則等待，預設等待超過10次則代表失敗。
    if checkExits('LD_HomeWudiAppIcon') == False:
        print('LD_HomeWudiAppIcon Fail')
        return

    #檢查試玩按鈕是否存在
    if checkExits('TrialPlay',True,30) == False:
        print('TrialPlay Fail')
        return

    #檢查廣告關閉按鈕是否存在
    #if checkExits('AdClose') == False:
    #    print('AdClose Fail')
    #    return

    #檢查公告關閉按鈕是否存在
    #if checkExits('Announce') == False:
    #    print('Announce Fail')
    #    return

    #檢查電子館大廳按鈕是否存在
    if checkExits('GoInToGameLobby') == False:
        print('GoInToGameLobby Fail')
        return


    nGameCnt = 0
    SwitchToGame('6001',nGameCnt,0,0)

    nGameCnt = 1
    SwitchToGame('6002',nGameCnt,0,0)

    nGameCnt = 2
    SwitchToGame('6003',nGameCnt,0,0)

    nGameCnt = 3
    SwitchToGame('6004',nGameCnt,0,1)

    nGameCnt = 4
    SwitchToGame('6005',nGameCnt,0,3)

    nGameCnt = 5
    SwitchToGame('6006',nGameCnt,0,0)

    nGameCnt = 6
    SwitchToGame('6008',nGameCnt,0,0)

    nGameCnt = 7
    SwitchToGame('6009',nGameCnt,0,0)

    nGameCnt = 8
    SwitchToGame('6010',nGameCnt,0,0)

    nGameCnt = 9
    SwitchToGame('6012',nGameCnt,0,0)

    nGameCnt = 10
    SwitchToGame('6013',nGameCnt,0,0)

    nGameCnt = 11
    SwitchToGame('6014',nGameCnt,0,0)

    nGameCnt = 12
    SwitchToGame('6015',nGameCnt,0,0)

    nGameCnt = 13
    SwitchToGame('6016',nGameCnt,0,0)

    nGameCnt = 14
    SwitchToGame('6017',nGameCnt,0,0)

    nGameCnt = 0
    SwitchToGame('6001',nGameCnt,0,0)


    return

ImageGrab.grab = partial(ImageGrab.grab, all_screens=True)
wb = openpyxl.load_workbook('D:/GoogleDrive/RegressionTest/RT_SwitchGame.xlsx')
localtime = time.localtime()
sheet = wb.create_sheet(str(time.strftime("LD_%y%m%d-%H-%M", localtime)),0)
nGameCnt = 0
if __name__ == "__main__":
    main()