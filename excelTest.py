import openpyxl
from openpyxl.styles import Font, PatternFill       # 載入 Font 和 PatternFill 模組
import datetime



# wb = openpyxl.load_workbook('RegressionTest.xlsx')
#wb = openpyxl.load_workbook('D:/GoogleDrive/RegressionTest/RegressionTest.xlsx')
#tonow = datetime.datetime.now()
#TodayDate = str(tonow.year) + str(tonow.month) + str(tonow.day)
#print(TodayDate)
sheet = wb.create_sheet(str(datetime.date.today()),0)


sheet.cell(1,1).value = "Game"
sheet.cell(1,2).value = "Table 1"
sheet.cell(1,3).value = "2"
sheet.cell(1,4).value = "3"
sheet.cell(1,5).value = "4"
sheet.cell(1,6).value = "5"
sheet.cell(1,7).value = "6"
sheet.cell(1,8).value = "7"
sheet.cell(1,9).value = "8"
sheet.cell(1,10).value = "9"
sheet.cell(1,11).value = "10"
sheet.cell(1,12).value = "11"
sheet.cell(1,13).value = "12"
sheet.cell(1,14).value = "13"
sheet.cell(1,15).value = "14"
sheet.cell(1,16).value = "15"
sheet.cell(1,17).value = "16"
sheet.cell(1,18).value = "17"
sheet.cell(1,19).value = "18"
sheet.cell(1,20).value = "19"
sheet.cell(1,21).value = "20"

sheet.cell(2,1).value = "6001"
sheet.cell(2,2).value = "OK"
sheet.cell(2,3).value = "X"
sheet.cell(2,3).fill = PatternFill(fill_type="solid", fgColor="FF0000")
sheet.cell(2,4).value = "OK"
sheet.cell(2,5).value = "OK"
sheet.cell(2,6).value = "OK"
sheet.cell(2,7).value = "OK"
sheet.cell(2,8).value = "OK"
sheet.cell(2,9).value = "OK"
sheet.cell(2,10).value = "OK"
sheet.cell(2,11).value = "OK"
sheet.cell(2,12).value = "OK"
sheet.cell(2,13).value = "OK"
sheet.cell(2,14).value = "OK"
sheet.cell(2,15).value = "OK"
sheet.cell(2,16).value = "OK"
sheet.cell(2,17).value = "OK"
sheet.cell(2,18).value = "OK"
sheet.cell(2,19).value = "OK"
sheet.cell(2,20).value = "OK"
sheet.cell(2,21).value = "OK"

#wb.save("D:/GoogleDrive/RegressionTest/RegressionTest.xlsx")
wb.save("RegressionTest_Save.xlsx")
