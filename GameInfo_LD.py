import pyautogui
import pyscreenshot as ImageGrab
import time
from PIL import ImageGrab
from functools import partial
from PIL import Image

import openpyxl
from openpyxl.styles import Font, PatternFill       # 載入 Font 和 PatternFill 模組
import datetime

import cv2
import numpy as np
from matplotlib import pyplot as plt
import imagehash
nGameCnt = 0

def checkInfo(PicName, Times = 15):
    nCheckTimes = 0
    iconAppCheck = False
    PicNamePath = 'D:/Selenium/auto_test/GameInfo/' + PicName + '.png'
    imgIcon = cv2.imread(PicNamePath)
    print(PicNamePath)
    iconW = imgIcon.shape[1]
    iconH = imgIcon.shape[0]

    while iconAppCheck == False:
        screen = pyautogui.screenshot()
        screen.save("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        screen = cv2.imread("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        res = cv2.matchTemplate(screen,imgIcon,1)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        #print(min_val, max_val, min_loc, max_loc)
        top_left = min_loc
        #print(top_left)

        bottom_right = (top_left[0] + iconW, top_left[1] + iconH)
        imgscreenIcon = ImageGrab.grab(bbox=(top_left[0], top_left[1], bottom_right[0], bottom_right[1]))
        #存檔確認比對圖案是否正確,此問題待解，希望可以不要再存檔重新讀取
        PicIconTempName = "D:/Selenium/auto_test/PicTemp/iconTemp.png"
        imgscreenIcon.save(PicIconTempName)

        hash1 = imagehash.average_hash(Image.open(PicNamePath))
        hash2 = imagehash.average_hash(Image.open(PicIconTempName))

        if hash1 == hash2:
            iconAppCheck = True
        else:
            print(nCheckTimes)
            nCheckTimes = nCheckTimes + 1
            time.sleep(1)
            if nCheckTimes >= Times:
                print('compare over times ' + PicName)
                return False
    return True

def checkExits(PicName,bClickIconFound = True, Times = 15, findGameIcon = False):
    nCheckTimes = 0
    iconAppCheck = False
    PicNamePath = 'D:/Selenium/auto_test/GameInfo/' + PicName + '.png'
    imgIcon = cv2.imread(PicNamePath)
    print(PicNamePath)
    iconW = imgIcon.shape[1]
    iconH = imgIcon.shape[0]

    while iconAppCheck == False:
        screen = pyautogui.screenshot()
        screen.save("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        screen = cv2.imread("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        res = cv2.matchTemplate(screen,imgIcon,1)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        #print(min_val, max_val, min_loc, max_loc)
        top_left = min_loc
        #print(top_left)

        bottom_right = (top_left[0] + iconW, top_left[1] + iconH)
        imgscreenIcon = ImageGrab.grab(bbox=(top_left[0], top_left[1], bottom_right[0], bottom_right[1]))
        #存檔確認比對圖案是否正確,此問題待解，希望可以不要再存檔重新讀取
        PicIconTempName = "D:/Selenium/auto_test/PicTemp/iconTemp.png"
        imgscreenIcon.save(PicIconTempName)
        imgscreenIcon2 = cv2.imread(PicIconTempName)
        hsv_base = cv2.cvtColor(imgIcon, cv2.COLOR_BGR2HSV)
        hsv_test = cv2.cvtColor(imgscreenIcon2, cv2.COLOR_BGR2HSV)
        h_bins = 50
        s_bins = 60
        histSize = [h_bins, s_bins]
        h_ranges = [0, 180]
        s_ranges = [0, 256]
        ranges = h_ranges + s_ranges
        
        channels = [0, 1]
        hist_base = cv2.calcHist([hsv_base], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_base, hist_base, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        hist_test = cv2.calcHist([hsv_test], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_test, hist_test, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        
        compare_method = cv2.HISTCMP_CORREL
        base_test = cv2.compareHist(hist_base, hist_test, compare_method)

        print(PicName + ' Similarity = ', base_test)
        if base_test > 0.9:
            print('Found ' + PicName)
            if bClickIconFound == True:
                pyautogui.moveTo(top_left[0] + iconW/2,top_left[1] + iconH/2,0.5)
                pyautogui.click()
            iconAppCheck = True
        else:
            if findGameIcon == True:
                pyautogui.moveTo(2850,1100,0.5)
                pyautogui.mouseDown(2850,1100)
                pyautogui.moveTo(2850,650,2)
                time.sleep(0.3)
                pyautogui.mouseUp(2850,650)
            print(nCheckTimes)
            nCheckTimes = nCheckTimes + 1
            time.sleep(1)
            if nCheckTimes >= Times:
                print('compare over times ' + PicName)
                return False
    return True

def checkGameInfo(GameName,WritePos):
    #檢查是否到遊戲館大廳
    if checkExits('Trial_MenuMore',False) == False:
        print('Trial_MenuMore Fail')
    
    time.sleep(1)
    #檢查 game Icon是否存在
    if checkExits(GameName,True,20,True) == False:
        print(GameName + ' check Fail')

    #檢查進入說明按鈕是否存在
    if checkExits('InfoBook') == False:
        print('InfoBook Fail')

    #檢查說明頁面是否正確
    InfoGameName = GameName + '_Info'
    if checkInfo(InfoGameName,10) == False:
        sheet.cell(WritePos + 2,2).value = "X"
        print(InfoGameName + ' Fail')
    else:
        sheet.cell(WritePos + 2,2).value = "OK"
        print("Found Info")

    #檢查進入關閉說明按鈕是否存在
    if checkExits('Info_Close') == False:
        print('Info_Close Fail')

    #檢查進入遊戲按鈕是否存在
    if checkExits('IntoDesk') == False:
        print('IntoDesk Fail')

    #檢查進入開啟選單按鈕是否存在
    if checkExits('MenuForOpenInfo') == False:
        print('MenuForOpenInfo Fail')

    #檢查進入說明按鈕是否存在
    if checkExits('InfoInDesk') == False:
        print('InfoInDesk Fail')

    #檢查說明頁面是否正確
    if checkInfo(InfoGameName,10) == False:
        sheet.cell(WritePos + 2,3).value = "X"
        print(InfoGameName + ' Fail')
    else:
        sheet.cell(WritePos + 2,3).value = "OK"
        print("Found Info")

    #檢查進入關閉說明按鈕是否存在
    if checkExits('Info_Close') == False:
        print('Info_Close Fail')

    #檢查進入說明按鈕是否存在
    if checkExits('TableDetail',False) == False:
        print('InfoInDesk Fail')

    #檢查可以入桌的桌號
    LoopNum = 0
    DeskNum = 0
    StartX = 2666
    StartY = 488
    FinalLoopStartY = 560
    XDistance = 117
    YDistance = 150
    ColumnItemNum = 5
    bFoundIntoDesk = False

    while LoopNum < 5:
        while DeskNum <= 19:
            ClickX = StartX + (DeskNum % ColumnItemNum) * XDistance
            if LoopNum == 4:
                ClickY = FinalLoopStartY + (DeskNum // ColumnItemNum) * YDistance
            else:
                ClickY = StartY + (DeskNum // ColumnItemNum) * YDistance
            pyautogui.moveTo(ClickX,ClickY)
            pyautogui.click()
            time.sleep(0.3)
            if checkExits('IntoGame',True,1) == False:
                print('IntoGame Fail')
            else:
                bFoundIntoDesk = True
                break
            DeskNum = DeskNum + 1
        if LoopNum == 4:
            break
        if bFoundIntoDesk == True:
            break
        DeskNum = 0        
        pyautogui.moveTo(2850,1000,0.5)
        pyautogui.mouseDown(2850,1000)
        pyautogui.moveTo(2850,413,2)
        time.sleep(0.3)
        pyautogui.mouseUp(2850,413)
        LoopNum = LoopNum + 1

    time.sleep(2)
    #檢查進入開啟選單按鈕是否存在
    if checkExits('Trial_MenuMore') == False:
        print('Trial_MenuMore Fail')

    #檢查進入說明按鈕是否存在
    if checkExits('InfoInDesk') == False:
        print('InfoInDesk Fail')

    #檢查說明頁面是否正確
    if checkInfo(InfoGameName,10) == False:
        sheet.cell(WritePos + 2,4).value = "X"
        print(InfoGameName + ' Fail')
    else:
        sheet.cell(WritePos + 2,4).value = "OK"
        print("Found Info")

    #檢查進入關閉說明按鈕是否存在
    if checkExits('Info_Close') == False:
        print('Info_Close Fail')

    #檢查進入上一頁按鈕是否存在
    if checkExits('BackToUpLevel') == False:
        print('BackToUpLevel Fail')

    #檢查進入說明按鈕是否存在
    if checkExits('TableDetail',False) == False:
        print('InfoInDesk Fail')

    #檢查進入上一頁按鈕是否存在
    if checkExits('BackToUpLevel') == False:
        print('BackToUpLevel Fail')

    sheet.cell(WritePos + 2,1).value = GameName
    wb.save("D:/GoogleDrive/RegressionTest/RT_GameInfo.xlsx")


    #檢查是否到遊戲館大廳
    #if checkExits('CheckInDesk',False) == False:
    #    print('CheckInDesk Fail')


    return True


def checkGameInfo_SP(GameName,WritePos):
    #檢查是否到遊戲館大廳
    if checkExits('Trial_MenuMore',False) == False:
        print('Trial_MenuMore Fail')
    
    time.sleep(1)
    #檢查 game Icon是否存在
    if checkExits(GameName,True,20,True) == False:
        print(GameName + ' check Fail')

    #檢查進入說明按鈕是否存在
    if checkExits('InfoBook') == False:
        print('InfoBook Fail')

    #檢查說明頁面是否正確
    InfoGameName = GameName + '_Info'
    if checkInfo(InfoGameName,10) == False:
        sheet.cell(WritePos + 2,2).value = "X"
        print(InfoGameName + ' Fail')
    else:
        sheet.cell(WritePos + 2,2).value = "OK"
        print("Found Info")

    #檢查進入關閉說明按鈕是否存在
    if checkExits('Info_Close') == False:
        print('Info_Close Fail')

    #檢查進入遊戲按鈕是否存在
    if checkExits('Into7001') == False:
        print('Into7001 Fail')

    sheet.cell(WritePos + 2,3).value = "Other"

    #很奇怪，會判斷錯誤，只好先sleep 5 秒
    time.sleep(5)

    #檢查進入是否進入遊戲
    if checkExits('checkInGame',False) == False:
        print('checkInGame Fail')

    #檢查進入開啟選單按鈕是否存在
    if checkExits('Trial_MenuMore') == False:
        print('Trial_MenuMore Fail')

    #檢查進入說明按鈕是否存在
    if checkExits('InfoInDesk') == False:
        print('InfoInDesk Fail')

    #檢查說明頁面是否正確
    if checkInfo(InfoGameName,10) == False:
        sheet.cell(WritePos + 2,4).value = "X"
        print(InfoGameName + ' Fail')
    else:
        sheet.cell(WritePos + 2,4).value = "OK"
        print("Found Info")

    #檢查進入關閉說明按鈕是否存在
    if checkExits('Info_Close') == False:
        print('Info_Close Fail')

    #檢查進入上一頁按鈕是否存在
    if checkExits('BackToUpLevel') == False:
        print('BackToUpLevel Fail')

    sheet.cell(WritePos + 2,1).value = GameName
    wb.save("D:/GoogleDrive/RegressionTest/RT_GameInfo.xlsx")

    return True
   
def main():

    sheet.cell(1,1).value = "Game"
    sheet.cell(1,2).value = "Lobby"
    sheet.cell(1,3).value = "Desk"
    sheet.cell(1,4).value = "Game"

    wb.save("D:/GoogleDrive/RegressionTest/RT_GameInfo.xlsx")

    #雷電 L_ForAutoTest
    pyautogui.moveTo(4191,46,0.5)
    pyautogui.doubleClick()
    
    #檢查模擬器桌面icon是否存在，若是則點擊進去，若否，則等待，預設等待超過10次則代表失敗。
    if checkExits('LD_HomeWudiAppIcon') == False:
        print('LD_HomeWudiAppIcon Fail')
        return

    #檢查試玩按鈕是否存在
    if checkExits('TrialPlay',True,30) == False:
        print('TrialPlay Fail')
        return

    #檢查廣告關閉按鈕是否存在
    if checkExits('AdClose') == False:
        print('AdClose Fail')
        return

    #檢查公告關閉按鈕是否存在
    #if checkExits('Announce') == False:
    #    print('Announce Fail')
    #    return

    #檢查電子館大廳按鈕是否存在
    if checkExits('GoInToGameLobby') == False:
        print('GoInToGameLobby Fail')
        return

    nGameCnt = 0
    checkGameInfo('6001',nGameCnt)
    
    nGameCnt = nGameCnt + 1
    checkGameInfo('6002',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6003',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6004',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6005',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6006',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6008',nGameCnt)
    
    nGameCnt = nGameCnt + 1
    checkGameInfo('6009',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6010',nGameCnt)
    
    nGameCnt = nGameCnt + 1
    checkGameInfo('6012',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6013',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6014',nGameCnt)

    nGameCnt = nGameCnt + 1
    checkGameInfo('6015',nGameCnt)
    
    nGameCnt = nGameCnt + 1
    checkGameInfo('6016',nGameCnt)
    
    nGameCnt = nGameCnt + 1
    checkGameInfo('6017',nGameCnt)
    
    nGameCnt = nGameCnt + 1
    checkGameInfo_SP('7001',nGameCnt)


    return

ImageGrab.grab = partial(ImageGrab.grab, all_screens=True)
wb = openpyxl.load_workbook('D:/GoogleDrive/RegressionTest/RT_GameInfo.xlsx')
localtime = time.localtime()
sheet = wb.create_sheet(str(time.strftime("LD_%y%m%d-%H-%M", localtime)),0)
nGameCnt = 0
if __name__ == "__main__":
    main()