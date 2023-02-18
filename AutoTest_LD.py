import pyautogui
import pyscreenshot as ImageGrab
import time
from PIL import ImageGrab
from functools import partial

import openpyxl
from openpyxl.styles import Font, PatternFill       # 載入 Font 和 PatternFill 模組
import datetime

import cv2
import numpy as np
from matplotlib import pyplot as plt
import pytesseract


def checkExits(PicName, Times = 15, findGameIcon = False):
    nCheckTimes = 0
    iconAppCheck = False
    PicNamePath = 'D:/Selenium/auto_test/PicForCompare/' + PicName + '.png'
    imgIcon = cv2.imread(PicNamePath)
    print(PicNamePath)
    iconW = imgIcon.shape[1]
    iconH = imgIcon.shape[0]

    while iconAppCheck == False:
        screen = pyautogui.screenshot()
        screen.save("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        screen = cv2.imread("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
        res = cv2.matchTemplate(screen,imgIcon,1)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        #print(min_val, max_val, min_loc, max_loc)
        top_left = min_loc
        #print(top_left)

        bottom_right = (top_left[0] + iconW, top_left[1] + iconH)
        imgscreenIcon = ImageGrab.grab(bbox=(top_left[0], top_left[1], bottom_right[0], bottom_right[1]))
        #存檔確認比對圖案是否正確,此問題待解，希望可以不要再存檔重新讀取
        PicIconTempName = "D:/Selenium/auto_test/PicTemp/iconTemp.png"
        imgscreenIcon.save(PicIconTempName)
        imgscreenIcon2 = cv2.imread(PicIconTempName)
        hsv_base = cv2.cvtColor(imgIcon, cv2.COLOR_BGR2HSV)
        hsv_test = cv2.cvtColor(imgscreenIcon2, cv2.COLOR_BGR2HSV)
        h_bins = 50
        s_bins = 60
        histSize = [h_bins, s_bins]
        h_ranges = [0, 180]
        s_ranges = [0, 256]
        ranges = h_ranges + s_ranges
        
        channels = [0, 1]
        hist_base = cv2.calcHist([hsv_base], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_base, hist_base, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        hist_test = cv2.calcHist([hsv_test], channels, None, histSize, ranges, accumulate=False)
        cv2.normalize(hist_test, hist_test, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        
        compare_method = cv2.HISTCMP_CORREL
        base_test = cv2.compareHist(hist_base, hist_test, compare_method)

        print(PicName + ' Similarity = ', base_test)
        if base_test > 0.9:
            print('Found ' + PicName)
            pyautogui.moveTo(top_left[0] + iconW/2,top_left[1] + iconH/2,0.5)
            pyautogui.click()
            iconAppCheck = True
        else:
            if findGameIcon == True:
                pyautogui.moveTo(2850,1100,0.5)
                pyautogui.mouseDown(2850,1100)
                pyautogui.moveTo(2850,650,2)
                time.sleep(0.3)
                pyautogui.mouseUp(2850,650)
            print(nCheckTimes)
            nCheckTimes = nCheckTimes + 1
            time.sleep(1)
            if nCheckTimes >= Times:
                print('compare over times ' + PicName)
                return False
    return True

def checkTableExits(tableNum,clickX,clickY):
    time.sleep(2)
    PicNamePath = 'D:/Selenium/auto_test/PicForCompare/TableClickCompare.png'
    imgIcon = cv2.imread(PicNamePath)

    iconW = imgIcon.shape[1]
    iconH = imgIcon.shape[0]
    screen = pyautogui.screenshot()
    screen.save("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
    screen = cv2.imread("D:/Selenium/auto_test/PicTemp/ScreenTemp.png")
    res = cv2.matchTemplate(screen,imgIcon,1)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
    #print(min_val, max_val, min_loc, max_loc)
    top_left = min_loc
    #print(top_left)

    bottom_right = (top_left[0] + iconW, top_left[1] + iconH)
    pyautogui.moveTo(clickX,clickY,0.2)
    pyautogui.click()
    time.sleep(0.1)
    imgscreenIcon = ImageGrab.grab(bbox=(top_left[0], top_left[1], bottom_right[0], bottom_right[1]))
    #存檔確認比對圖案是否正確,此問題待解，希望可以不要再存檔重新讀取
    PicIconTempName = "D:/Selenium/auto_test/PicTemp/iconTemp.png"
    imgscreenIcon.save(PicIconTempName)
    imgscreenIcon2 = cv2.imread(PicIconTempName)
    hsv_base = cv2.cvtColor(imgIcon, cv2.COLOR_BGR2HSV)
    hsv_test = cv2.cvtColor(imgscreenIcon2, cv2.COLOR_BGR2HSV)
    h_bins = 50
    s_bins = 60
    histSize = [h_bins, s_bins]
    h_ranges = [0, 180]
    s_ranges = [0, 256]
    ranges = h_ranges + s_ranges
        
    channels = [0, 1]
    hist_base = cv2.calcHist([hsv_base], channels, None, histSize, ranges, accumulate=False)
    cv2.normalize(hist_base, hist_base, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
    hist_test = cv2.calcHist([hsv_test], channels, None, histSize, ranges, accumulate=False)
    cv2.normalize(hist_test, hist_test, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        
    compare_method = cv2.HISTCMP_CORREL
    base_test = cv2.compareHist(hist_base, hist_test, compare_method)

    print('Table '+ str(tableNum) + ' Similarity = ', base_test)
    if base_test < 0.9:
        sheet.cell(nGameCnt + 2,tableNum + 2).value = "OK"
        print("table " + str(tableNum) + " actived!!!!")
    else:
        sheet.cell(nGameCnt + 2,tableNum + 2).value = "X"
        sheet.cell(nGameCnt + 2,tableNum + 2).fill = PatternFill(fill_type="solid", fgColor="FF0000")
        print("table " + str(tableNum) + " untouchedble~~~")
        return False
    return True

def checkDeskExits():
    #第一排
    tableCnt = 1
    checkTableExits(tableCnt,2665,485)
  
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2780,485)
    
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2894,485)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3009,485)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3124,485)

    #第2排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2665,622)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2780,622)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2894,622)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3009,622)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3124,622)

    #第3排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2665,766)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2780,766)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2894,766)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3009,766)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3124,766)

    #第4排
    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2665,911)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2780,911)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,2894,911)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3009,911)

    tableCnt = tableCnt + 1
    checkTableExits(tableCnt,3124,911)
    return

def checkGameExits(GameName):
    #檢查是否到遊戲館大廳
    if checkExits('Trial_MenuMore') == False:
        print('Trial_MenuMore Fail')
    #關閉選單
    pyautogui.click()
    
    #檢查 game Icon是否存在
    if checkExits(GameName,20,True) == False:
        print(GameName + ' check Fail')

    #檢查進入遊戲按鈕是否存在
    if checkExits('IntoDesk') == False:
        print('IntoDesk Fail')

    #檢查 遊戲內橫幅圖是否存在，用以確認真的進到遊戲選桌畫面中
    #gameBaner = GameName + '_Banner'
    #if checkExits(gameBaner) == False:
    #    print(gameBaner +' check Fail')

    #檢查桌詳細資訊按鈕是否存在，順便確認是否真的進到遊戲選桌畫面中
    if checkExits('TableDetail') == False:
        print('TableDetail check Fail')

    #檢查桌詳細資訊是否可以打開，順便確認是否真的進到遊戲選桌畫面中
    if checkExits('TableDetailInfo') == True:
        sheet.cell(nGameCnt + 2,2).value = 'OK'
    else:
        sheet.cell(nGameCnt + 2,2).value = 'Fail'
        print('TableDetailInfo check Fail')
    
    #關閉桌詳細資訊
    if checkExits('TableDetail') == False:
        print('TableDetail check Fail')


    sheet.cell(nGameCnt + 2,1).value = GameName
    #檢查 遊戲內 各桌是否可以按
    checkDeskExits()
    
    wb.save("D:/GoogleDrive/RegressionTest/RT_DeskCheck.xlsx")


    #檢查離開遊戲按鈕是否存在
    if checkExits('BacktoGameLobby') == False:
        print('BacktoGameLobby Fail')

    return

def main():
    ImageGrab.grab = partial(ImageGrab.grab, all_screens=True)

    sheet.cell(1,1).value = "Game"
    sheet.cell(1,2).value = "Detail Info"
    sheet.cell(1,3).value = "Table 1"
    sheet.cell(1,4).value = "2"
    sheet.cell(1,5).value = "3"
    sheet.cell(1,6).value = "4"
    sheet.cell(1,7).value = "5"
    sheet.cell(1,8).value = "6"
    sheet.cell(1,9).value = "7"
    sheet.cell(1,10).value = "8"
    sheet.cell(1,11).value = "9"
    sheet.cell(1,12).value = "10"
    sheet.cell(1,13).value = "11"
    sheet.cell(1,14).value = "12"
    sheet.cell(1,15).value = "13"
    sheet.cell(1,16).value = "14"
    sheet.cell(1,17).value = "15"
    sheet.cell(1,18).value = "16"
    sheet.cell(1,19).value = "17"
    sheet.cell(1,20).value = "18"
    sheet.cell(1,21).value = "19"
    sheet.cell(1,22).value = "20"

    wb.save("D:/GoogleDrive/RegressionTest/RT_DeskCheck.xlsx")

    #雷電 L_ForAutoTest
    pyautogui.moveTo(4191,46,0.5)
    pyautogui.doubleClick()
    
    #檢查模擬器桌面icon是否存在，若是則點擊進去，若否，則等待，預設等待超過10次則代表失敗。
    if checkExits('LD_HomeWudiAppIcon') == False:
        print('LD_HomeWudiAppIcon Fail')
        return

    #檢查試玩按鈕是否存在
    if checkExits('TrialPlay', 30) == False:
        print('TrialPlay Fail')
        return

    #這次版本沒有廣告
    #檢查廣告關閉按鈕是否存在
    if checkExits('AdClose') == False:
        print('AdClose Fail')
        return

    #這次版本沒有廣告
    #檢查公告關閉按鈕是否存在
    if checkExits('Announce') == False:
        print('Announce Fail')
        return

    #檢查電子館大廳按鈕是否存在
    if checkExits('GoInToGameLobby') == False:
        print('GoInToGameLobby Fail')
        return

    nGameCnt = 0
    checkGameExits('6001')

    nGameCnt = nGameCnt + 1
    checkGameExits('6002')

    nGameCnt = nGameCnt + 1
    checkGameExits('6003')

    nGameCnt = nGameCnt + 1
    checkGameExits('6004')

    nGameCnt = nGameCnt + 1
    checkGameExits('6005')

    nGameCnt = nGameCnt + 1
    checkGameExits('6006')

    nGameCnt = nGameCnt + 1
    checkGameExits('6008')
    
    nGameCnt = nGameCnt + 1
    checkGameExits('6009')

    nGameCnt = nGameCnt + 1
    checkGameExits('6010')
    
    nGameCnt = nGameCnt + 1
    checkGameExits('6012')

    nGameCnt = nGameCnt + 1
    checkGameExits('6013')

    nGameCnt = nGameCnt + 1
    checkGameExits('6014')

    nGameCnt = nGameCnt + 1
    checkGameExits('6015')

    nGameCnt = nGameCnt + 1
    checkGameExits('6016')
    
    nGameCnt = nGameCnt + 1
    checkGameExits('6017')

wb = openpyxl.load_workbook('D:/GoogleDrive/RegressionTest/RT_DeskCheck.xlsx')
localtime = time.localtime()
sheet = wb.create_sheet(str(time.strftime("LD_%y%m%d-%H-%M", localtime)),0)
nGameCnt = 0
if __name__ == "__main__":
    main()